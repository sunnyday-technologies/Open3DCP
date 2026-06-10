"""Readers that turn external dataset files into normalized *source records*.

A source record is a dict mapping a fully-qualified source field path
(e.g. "material_batches.cement_content_kg_m3", "tests.age_days",
"data.<quantity>.mean") to a value, plus a "_ctx" entry with conversion context
(total wet mass, units, etc.). The ingest engine maps these onto Open3DCP columns.
"""
from __future__ import annotations

import csv
import os
from typing import Any


# ---------------------------------------------------------------------------
# UCI "Concrete Compressive Strength" (Yeh 1998) -- flat CSV, kg/m3
# ---------------------------------------------------------------------------
_UCI_ALIASES = {
    "cement": "Cement", "blast furnace slag": "Blast Furnace Slag", "slag": "Blast Furnace Slag",
    "fly ash": "Fly Ash", "water": "Water", "superplasticizer": "Superplasticizer",
    "coarse aggregate": "Coarse Aggregate", "fine aggregate": "Fine Aggregate",
    "age": "Age", "concrete compressive strength": "Concrete compressive strength",
    "csmpa": "Concrete compressive strength",
}
# UCI components that sum to the full wet-mix mass (all kg/m3, incl. water).
_UCI_MASS_COMPONENTS = ["Cement", "Blast Furnace Slag", "Fly Ash", "Water",
                        "Superplasticizer", "Coarse Aggregate", "Fine Aggregate"]


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _canon_uci(header: str) -> str | None:
    h = header.strip().lower()
    # strip the "(component N)(unit)" decorations Yeh's CSV carries
    for key, canon in _UCI_ALIASES.items():
        if h.startswith(key):
            return canon
    return None


def read_uci_csv(path: str) -> list[dict[str, Any]]:
    records = []
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        header = next(reader)
        canon = [_canon_uci(h) for h in header]
        for raw in reader:
            if not any(c.strip() for c in raw):
                continue
            rec: dict[str, Any] = {}
            for col, val in zip(canon, raw):
                if col is None:
                    continue
                rec[col] = _num(val)
            total = sum(rec.get(c) or 0.0 for c in _UCI_MASS_COMPONENTS)
            binder = sum(rec.get(c) or 0.0 for c in ("Cement", "Blast Furnace Slag", "Fly Ash"))
            rec["_ctx"] = {
                "total_wet_mass_kg_m3": total if total > 0 else None,
                "total_wet_mass_is_complete": True,  # UCI reports every constituent incl. water
                "total_binder_kg_m3": binder if binder > 0 else None,
                "total_batched_mass_kg_m3": total if total > 0 else None,
                "source": "uci_yeh_1998",
            }
            records.append(rec)
    return records


# ---------------------------------------------------------------------------
# relational concrete database .xlsx template -- kg/m3
# ---------------------------------------------------------------------------
def _load_xlsx_tabs(path: str) -> dict[str, list[dict[str, Any]]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    tabs: dict[str, list[dict[str, Any]]] = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            tabs[ws.title] = []
            continue
        # the template uses a single header row for entity tabs; some tabs carry a
        # banner row above the header (reinforcement_components, data) -- detect the
        # header row as the first row whose first cell ends with "_id".
        header_idx = 0
        for i, r in enumerate(rows[:3]):
            first = (r[0] or "") if r else ""
            if isinstance(first, str) and first.strip().endswith("_id"):
                header_idx = i
                break
        header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
        body = []
        for r in rows[header_idx + 1:]:
            if not any(c is not None and str(c).strip() for c in r):
                continue
            body.append({h: v for h, v in zip(header, r) if h})
        tabs[ws.title] = body
    wb.close()
    return tabs


_BINDER_KG = ["cement_content_kg_m3", "silica_fume_content_kg_m3",
              "fly_ash_content_kg_m3", "slag_content_kg_m3", "metakaolin_content_kg_m3"]
# Every kg/m3 constituent that contributes mass to the fresh wet mix. The total wet mass
# (denominator of every mass-%) MUST include all of these -- omitting admixtures/SCMs
# biases the composition vector high. Aggregates and kg/m3 admixtures are mass-accountable.
_MASS_KG_FIELDS = _BINDER_KG + [
    "limestone_content_kg_m3", "fine_aggregate_content_kg_m3",
    "coarse_aggregate_content_kg_m3", "superplasticizer_content_kg_m3",
    "rheology_modifier_content_kg_m3",
]
# Constituents the source reports by VOLUME or as a dose, not a kg/m3 mass: their mass cannot
# be added to the wet-mass balance without a density. If any is present, the balance is not
# closed -> `complete=False` so the kg/m3 -> mass-% conversion is honestly flagged lossy
# (rather than reported "exact" against a denominator that silently omits them).
_UNACCOUNTED_FIELDS = [
    "water_reducer_content_ml_m3", "air_entrainment_content_ml_m3",
    "hydration_accelerator_content_ml_m3", "fiber_volume_fraction",
]


def _batch_total_wet_mass(batch: dict):
    """Estimate total wet-mix mass (kg/m3) for a source batch.

    total = Σ(all kg/m3 constituents) + water,  water = w/b * binder.
    Returns (total, is_complete, binder). `complete` is True only when the mass balance is
    actually closed: binder, w/b and aggregates are present AND no constituent is reported
    by volume/dose (ml/m3, fiber volume fraction) whose mass we cannot account for.
    """
    binder = sum(_num(batch.get(k)) or 0.0 for k in _BINDER_KG)
    wb = _num(batch.get("water_binder_ratio"))
    fine = _num(batch.get("fine_aggregate_content_kg_m3")) or 0.0
    coarse = _num(batch.get("coarse_aggregate_content_kg_m3")) or 0.0
    binder_out = binder if binder > 0 else None
    if binder <= 0 or wb is None:
        return (None, False, binder_out)
    water = wb * binder
    mass_kg = sum(_num(batch.get(k)) or 0.0 for k in _MASS_KG_FIELDS)
    total = mass_kg + water
    unaccounted = any(_num(batch.get(k)) is not None for k in _UNACCOUNTED_FIELDS)
    complete = binder > 0 and wb is not None and (fine + coarse) > 0 and not unaccounted
    return (total, complete, binder_out)


def read_relational_xlsx(path: str) -> list[dict[str, Any]]:
    """Denormalize the relational template into per-(batch x test) source records.

    Joins material_batches -> specimens -> tests -> data and pivots each `data` row
    (one quantity) under "data.<quantity>.{mean,std,units}". Fields that don't fit the
    flat grain are still attached (prefixed by tab) so the engine can route them to the
    triage sidecar -- nothing is dropped at read time.
    """
    tabs = _load_xlsx_tabs(path)
    batches = {b.get("batch_id"): b for b in tabs.get("material_batches", [])}
    specimens = tabs.get("specimens", [])
    tests_by_specimen: dict[Any, list] = {}
    for t in tabs.get("tests", []):
        tests_by_specimen.setdefault(t.get("specimen_id"), []).append(t)
    data_by_test: dict[Any, list] = {}
    for d in tabs.get("data", []):
        data_by_test.setdefault(d.get("test_id"), []).append(d)

    records: list[dict[str, Any]] = []
    # specimens drive the grain; if a specimen has no test, still emit the batch row.
    seen_specimens = set()
    for spec in specimens:
        seen_specimens.add(spec.get("specimen_id"))
        batch = batches.get(spec.get("batch_id"), {})
        total, complete, binder = _batch_total_wet_mass(batch)
        spec_tests = tests_by_specimen.get(spec.get("specimen_id"), [None])
        for test in spec_tests:
            rec: dict[str, Any] = {}
            for k, v in batch.items():
                rec[f"material_batches.{k}"] = v
            for k, v in spec.items():
                rec[f"specimens.{k}"] = v
            if test:
                for k, v in test.items():
                    rec[f"tests.{k}"] = v
                for d in data_by_test.get(test.get("test_id"), []):
                    q = d.get("quantity_reported")
                    if q:
                        # Attach data_type and file_name PER quantity (not once per test) so a
                        # test carrying both a scalar and a curve keeps each record's routing
                        # metadata -- the engine reads data.<q>.data_type to send curves/images
                        # to the right *_file column instead of dropping the descriptor.
                        rec[f"data.{q}.mean"] = d.get("quantity_reported_mean")
                        rec[f"data.{q}.std"] = d.get("quantity_reported_standard_deviation")
                        rec[f"data.{q}.units"] = d.get("units")
                        if d.get("data_type") is not None:
                            rec[f"data.{q}.data_type"] = d.get("data_type")
                        if d.get("file_name") is not None:
                            rec[f"data.{q}.file_name"] = d.get("file_name")
                    for k in ("number_of_specimens", "extraction_methods",
                              "curator_first_name", "curator_last_name"):
                        if d.get(k) is not None:
                            rec[f"data.{k}"] = d.get(k)
            rec["_ctx"] = {
                "total_wet_mass_kg_m3": total,
                "total_wet_mass_is_complete": complete,
                "total_binder_kg_m3": binder,
                "total_batched_mass_kg_m3": total,
                "source": "relational",
                "source_id": batch.get("source_id") or spec.get("source_id"),
            }
            records.append(rec)
    # batches with no specimen at all -> still emit so nothing is lost
    referenced = {s.get("batch_id") for s in specimens}
    for bid, batch in batches.items():
        if bid in referenced:
            continue
        total, complete, binder = _batch_total_wet_mass(batch)
        rec = {f"material_batches.{k}": v for k, v in batch.items()}
        rec["_ctx"] = {"total_wet_mass_kg_m3": total, "total_wet_mass_is_complete": complete,
                       "total_binder_kg_m3": binder, "total_batched_mass_kg_m3": total,
                       "source": "relational", "source_id": batch.get("source_id")}
        records.append(rec)
    return records


# ---------------------------------------------------------------------------
# generic flat kg/m3 mix table -- a curated CSV with canonical column tokens
# ---------------------------------------------------------------------------
# Many open datasets are a single flat table that reports every constituent in
# kg/m3 of a 1 m3 batch (UCI, Meta SustainableConcrete, ...).
# This reader consumes a curated excerpt whose headers are canonical tokens we
# control (so the reader is dataset-agnostic) and emits source records using the
# SAME field paths as the relational reader, so the shared relational crosswalk
# maps them. It computes the wet-mass balance from ALL kg/m3 constituents (the
# denominator of every mass-%), so kg/m3 -> mass-% is exact when the batch closes.
_FLAT_BINDERS = {  # canonical CSV column -> material_batches.<field>
    "cement_kg_m3": "cement_content_kg_m3",
    "fly_ash_kg_m3": "fly_ash_content_kg_m3",
    "slag_kg_m3": "slag_content_kg_m3",
    "silica_fume_kg_m3": "silica_fume_content_kg_m3",
    "metakaolin_kg_m3": "metakaolin_content_kg_m3",
    "limestone_kg_m3": "limestone_content_kg_m3",
    "nano_silica_kg_m3": "nano_silica_content_kg_m3",
    "mineral_powder_kg_m3": "mineral_powder_content_kg_m3",
}
_FLAT_OTHER_MASS = {  # kg/m3 constituents that add mass but are not binder
    "water_kg_m3": "water_content_kg_m3",
    "superplasticizer_kg_m3": "superplasticizer_content_kg_m3",
    "coarse_agg_kg_m3": "coarse_aggregate_content_kg_m3",
    "fine_agg_kg_m3": "fine_aggregate_content_kg_m3",
    "fiber_kg_m3": "fiber_content_kg_m3",
}
# binders that contribute to the binder total (denominator of w/b, total_binder)
_FLAT_BINDER_TOTAL = ["cement_kg_m3", "fly_ash_kg_m3", "slag_kg_m3", "silica_fume_kg_m3",
                      "metakaolin_kg_m3", "nano_silica_kg_m3"]
# data quantities: canonical CSV column -> (quantity_reported, units)
_FLAT_QUANTITIES = {
    "compressive_strength_mpa": ("compressive_strength", "MPa"),
    "flexural_strength_mpa": ("flexural_strength", "MPa"),
    "tensile_strength_mpa": ("tensile_strength", "MPa"),
    "splitting_tensile_mpa": ("splitting_tensile", "MPa"),
    "elastic_modulus_gpa": ("elastic_modulus", "GPa"),
    "slump_mm": ("slump", "mm"),
}
# Typical specific gravities for an absolute-volume YIELD check. The sum of constituent kg/m3 is the
# batched MASS, not a fresh density: if the constituents do not occupy ~1 m3 the source rows are
# design proportions (e.g. water added on top of a ~1 m3 dry recipe), not a yielded batch -- recorded
# as a provenance note, never stored as a "density".
_FLAT_SG = {
    "cement_kg_m3": 3.15, "fly_ash_kg_m3": 2.30, "slag_kg_m3": 2.90, "silica_fume_kg_m3": 2.20,
    "metakaolin_kg_m3": 2.50, "limestone_kg_m3": 2.70, "nano_silica_kg_m3": 2.20,
    "mineral_powder_kg_m3": 2.65, "water_kg_m3": 1.00, "superplasticizer_kg_m3": 1.07,
    "coarse_agg_kg_m3": 2.70, "fine_agg_kg_m3": 2.65, "fiber_kg_m3": 7.85,
}


def read_flat_csv(path: str) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for raw in csv.DictReader(fh):
            row = {k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in raw.items()}
            rec: dict[str, Any] = {}
            # mix-design constituents (kg/m3)
            for col, field in {**_FLAT_BINDERS, **_FLAT_OTHER_MASS}.items():
                val = _num(row.get(col))
                if val is not None:
                    rec[f"material_batches.{field}"] = val
            # categorical / identity batch fields
            for col, field in (("cement_type", "cement_type"), ("fly_ash_class", "fly_ash_class"),
                               ("material_class", "material_class"), ("fiber_type", "fiber_type"),
                               ("printable", "printable")):
                v = row.get(col)
                if v not in (None, ""):
                    rec[f"material_batches.{field}"] = v
            # default cement type to ASTM C150 Type I only when cement is present, and FLAG it as an
            # assumption (the source stated no type) so the fidelity scorer counts it -- not silently exact.
            if rec.get("material_batches.cement_content_kg_m3") and "material_batches.cement_type" not in rec:
                rec["material_batches.cement_type"] = "ASTM_C150_Type_I"
                rec.setdefault("_assumed_fields", set()).add("material_batches.cement_type")
            for col, field in (("fiber_length_mm", "fiber_length_mm"),
                               ("fiber_diameter_mm", "fiber_diameter_mm"),
                               ("max_agg_size_mm", "max_aggregate_size_mm"),
                               ("embodied_carbon_kg_co2_m3", "embodied_carbon_kg_co2_m3")):
                val = _num(row.get(col))
                if val is not None:
                    rec[f"material_batches.{field}"] = val
            # water/binder ratio (derived; both the water column and the ratio are kept)
            binder = sum(_num(row.get(c)) or 0.0 for c in _FLAT_BINDER_TOTAL)
            water = _num(row.get("water_kg_m3"))
            if binder > 0 and water is not None:
                rec["material_batches.water_binder_ratio"] = round(water / binder, 4)
            # specimen / test conditions
            if row.get("specimen_geometry"):
                rec["specimens.specimen_geometry"] = row["specimen_geometry"]
            if row.get("test_method"):
                rec["tests.test_type"] = row["test_method"]
            for col, field in (("age_days", "tests.age_days"),
                               ("curing_temp_c", "tests.initial_env_temperature_C"),
                               ("curing_humidity_pct", "tests.initial_env_relative_humidity_percent")):
                val = _num(row.get(col))
                if val is not None:
                    rec[field] = val
            # measured quantities -> data.<q>.{mean,std,units}
            for col, (q, unit) in _FLAT_QUANTITIES.items():
                mean = _num(row.get(col))
                if mean is None:
                    continue
                rec[f"data.{q}.mean"] = mean
                rec[f"data.{q}.units"] = unit
                if q == "compressive_strength":
                    std = _num(row.get("compressive_strength_std_mpa"))
                    if std is not None:
                        rec[f"data.{q}.std"] = std
            n = _num(row.get("n_specimens"))
            if n is not None:
                rec["data.number_of_specimens"] = int(n)
            rec["data.extraction_methods"] = row.get("extraction_method") or "direct"
            # wet-mass balance: every constituent is a kg/m3 mass, so the total closes
            mass_fields = list(_FLAT_BINDERS) + list(_FLAT_OTHER_MASS)
            total = sum(_num(row.get(c)) or 0.0 for c in mass_fields)
            # The wet-mass total only "closes" (and the kg/m3 -> mass-% projection is only exact) when
            # water is actually present; a binder+aggregate row with no water would otherwise be scored
            # against a waterless denominator and mis-rated EXACT. Mirror the relational guard.
            complete = binder > 0 and total > 0 and water is not None
            # Absolute-volume YIELD check: `total` is the batched MASS, not a fresh density. If the
            # constituents do not occupy ~1 m3, the source rows are design proportions (commonly: water
            # dosed on top of a ~1 m3 dry recipe), not a yielded batch. Record the true derived density
            # and the over/under-yield as a provenance note -- never store `total` as a "density".
            note = None
            abs_vol = sum((_num(row.get(c)) or 0.0) / _FLAT_SG[c]
                          for c in mass_fields if c in _FLAT_SG and _num(row.get(c)))
            if complete and abs_vol > 0:
                yield_m3 = abs_vol / 1000.0
                if abs(yield_m3 - 1.0) > 0.03:
                    note = (f"design proportions: absolute-volume yield {yield_m3:.3f} m3 "
                            f"({(yield_m3 - 1) * 100:+.0f}% vs a closed 1 m3 batch) -> derived fresh "
                            f"density ~{total / yield_m3:.0f} kg/m3; total_batched_mass_kg_m3 is the "
                            f"batched mass, not a measured density")
            rec["_ctx"] = {
                "total_wet_mass_kg_m3": total if total > 0 else None,
                "total_wet_mass_is_complete": complete,
                "total_binder_kg_m3": binder if binder > 0 else None,
                "total_batched_mass_kg_m3": total if total > 0 else None,
                "provenance_notes": note,
                "source": "flat",
            }
            records.append(rec)
    return records


def detect_and_read(path: str, kind: str | None = None) -> tuple[str, list[dict[str, Any]]]:
    """Dispatch on file type / explicit kind. Returns (kind, records)."""
    ext = os.path.splitext(path)[1].lower()
    if kind is None:
        kind = "relational" if ext in (".xlsx", ".xlsm") else "uci"
    if kind == "relational":
        return kind, read_relational_xlsx(path)
    if kind == "uci":
        return kind, read_uci_csv(path)
    if kind == "flat":
        return kind, read_flat_csv(path)
    raise ValueError(f"unknown source kind {kind!r}")
