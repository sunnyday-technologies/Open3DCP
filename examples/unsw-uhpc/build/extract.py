#!/usr/bin/env python3
"""Reproduce the UNSW Global UHPC dataset -> Open3DCP curated excerpt.

Source (CC BY 4.0), NOT re-hosted here -- download it from Mendeley Data and run:

    python build/extract.py "/path/to/UHPC Dataset  (Version-2).xlsx"

    Malik, U. J.; Lee, C. K.; Mohotti, D.; Mo, H. "A global dataset of UHPC mix designs..."
    Mendeley Data, V2, doi:10.17632/czb7ww5pkz.2  (CC BY 4.0)

The source `UHPC Dataset` tab is a single flat kg/m3 table (2,188 mixes) with a 3-row hierarchical
header. This script reads the fixed constituent/strength columns by position, selects a seven-mix
excerpt spanning the UHPC design space (silica-fume matrix; fly-ash / slag / metakaolin / quartz-powder
SCMs; with and without steel fibre; ~81-281 MPa (28-day 130-281); some carrying flexural, elastic-modulus and splitting-
tensile results), and writes the canonical flat CSV that `open3dcp-ingest --kind flat` consumes. Each
(mix x reported compressive age) becomes one row; 28-day flexural / modulus / splitting are attached to
the latest-age row. Values are copied unchanged (all native kg/m3, MPa, GPa); nothing is synthesized.
"""
import csv
import os
import sys

# Fixed (Mix-ID, rounded cement kg/m3) selectors -> first matching source row. The pair disambiguates
# repeated Mix-IDs across the 168 source studies.
SELECT = [
    ("F0", 935),       # silica-fume + quartz matrix, NO fibre; modulus + splitting tensile (199 MPa)
    ("F40", 561),      # 40% fly-ash replacement; modulus + splitting (153 MPa) -- the SCM-vs-strength trend
    ("S10-F10", 839),  # straight steel fibre + fly ash; flexural MOR; 7 & 28 d (135 MPa)
    ("S10-M10", 839),  # straight steel fibre + metakaolin; flexural MOR; 7 & 28 d (130 MPa)
    ("B-S0", 940),     # steel fibre, plain SF matrix; 200 MPa @ 28 d (219 @ 90 d)
    ("B-S40", 564),    # steel fibre + 40% GGBFS; 200 MPa @ 28 d (225 @ 90 d)
    ("G10F10", 664),   # steel fibre + GGBFS + fly ash (quaternary); 281 MPa
]

# fixed column positions in the 'UHPC Dataset' tab (decoded from its 3-row header)
IX = dict(mix=0, cement=1, sf=4, fa=5, quartz=8, mk=11, ggbfs=12, slag=13, nanosi=18, sand=21,
          fibtype=24, fibamt=25, fiblen=26, fibdia=27, water=36, sp=38,
          cs1=48, cs3=49, cs7=50, cs14=51, cs28=53, cs90=55, em=57, splittens=59, mor=68)
AGES = [("cs1", 1), ("cs3", 3), ("cs7", 7), ("cs14", 14), ("cs28", 28), ("cs90", 90)]

COLS = ["mix_id", "material_class", "cement_kg_m3", "silica_fume_kg_m3", "fly_ash_kg_m3",
        "slag_kg_m3", "metakaolin_kg_m3", "nano_silica_kg_m3", "mineral_powder_kg_m3",
        "water_kg_m3", "superplasticizer_kg_m3", "fine_agg_kg_m3",
        "fiber_type", "fiber_kg_m3", "fiber_length_mm", "fiber_diameter_mm",
        "age_days", "compressive_strength_mpa", "flexural_strength_mpa",
        "splitting_tensile_mpa", "elastic_modulus_gpa"]


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def fiber_token(raw):
    s = (raw or "").lower()
    if "steel" in s:
        return "steel"
    if "pva" in s or "polyvinyl" in s:
        return "pva"
    if "glass" in s:
        return "glass"
    if "poly" in s or "pp" in s:
        return "polypropylene"
    return ""


def main(src):
    from openpyxl import load_workbook
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb["UHPC Dataset "]
    data = list(ws.iter_rows(values_only=True))[3:]
    wb.close()
    chosen = []
    for mix_id, cem in SELECT:
        for r in data:
            if str(r[IX["mix"]]).strip() == mix_id and round(num(r[IX["cement"]]) or -1) == cem:
                chosen.append(r)
                break
        else:
            raise SystemExit(f"selector not found: {mix_id} / cement~{cem}")

    out = []
    for r in chosen:
        g = lambda k: num(r[IX[k]])
        slag = (g("ggbfs") or 0) + (g("slag") or 0)
        ftype = fiber_token(r[IX["fibtype"]])
        famt = g("fibamt")
        base = {
            "mix_id": str(r[IX["mix"]]).strip(), "material_class": "UHPC",
            "cement_kg_m3": g("cement"), "silica_fume_kg_m3": g("sf"), "fly_ash_kg_m3": g("fa"),
            "slag_kg_m3": slag or None, "metakaolin_kg_m3": g("mk"), "nano_silica_kg_m3": g("nanosi"),
            "mineral_powder_kg_m3": g("quartz"),  # quartz powder -> generic mineral filler
            "water_kg_m3": g("water"), "superplasticizer_kg_m3": g("sp"), "fine_agg_kg_m3": g("sand"),
            "fiber_type": ftype or None,
            "fiber_kg_m3": famt if (ftype and famt) else None,
            "fiber_length_mm": g("fiblen") if ftype else None,
            "fiber_diameter_mm": g("fibdia") if ftype else None,
        }
        ages = [(d, g(k)) for k, d in AGES if g(k) is not None]
        if not ages:
            continue
        latest = max(d for d, _ in ages)
        for d, cs in ages:
            row = dict(base)
            row["age_days"] = d
            row["compressive_strength_mpa"] = cs
            if d == latest:  # attach 28-d-style mechanical extras to the latest-age row
                row["flexural_strength_mpa"] = g("mor")
                row["splitting_tensile_mpa"] = g("splittens")
                row["elastic_modulus_gpa"] = g("em")
            out.append(row)

    dest = os.path.join(os.path.dirname(os.path.abspath(__file__)), "unsw-uhpc.csv")
    with open(dest, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=COLS)
        w.writeheader()
        for d in out:
            w.writerow({k: ("" if d.get(k) in (None, "") else d.get(k)) for k in COLS})
    print(f"wrote {len(out)} rows ({len(chosen)} mixes) -> {dest}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit('usage: python build/extract.py "/path/to/UHPC Dataset  (Version-2).xlsx"')
    main(sys.argv[1])
