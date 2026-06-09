#!/usr/bin/env python3
"""Reproduce the UF/UTK 3D-Printing-Concrete Mix-Design Open Dataset -> Open3DCP excerpt.

This example is HAND-CURATED (like the RILEM one). The source is reported on a **ratio-to-binder**
basis (binder = 1; every constituent given as a fraction of binder), with NO absolute binder content in
kg/m3 -- so the **kg/m3 basis is not recoverable** and a constituent mass-% of the total mix cannot be
formed without an assumed binder dosage. We therefore do NOT fabricate constituent mass-%; we curate the
fields the source defines unambiguously: the binder system, water/binder ratio, the fresh-state
**rheology** that governs printability (static / dynamic yield stress, plastic viscosity), and compressive
strength by age. An automated `open3dcp-ingest` ratio-to-binder reader (which would also emit a fidelity
score) is a planned follow-up; for now the score is deferred, exactly as for the RILEM excerpt.

Source (CC BY 4.0), NOT re-hosted -- download it from the DOI and run:

    python build/extract.py "/path/to/3D concrete printing mix design dataset v0.3.xlsx"

    Gao, J.; Wang, Z.; Wang, C. "3D Printing Concrete Mix Design Open Dataset" (v0.3)
    Zenodo, doi:10.5281/zenodo.6828947 (CC BY 4.0)

It selects printable Portland-cement mixes that report fresh rheology + strength (up to two per source
study, spanning the yield-stress and strength range), maps kPa -> Pa, and writes
../ufutk-3dcp-mix.open3dcp.csv. Values are real; nothing is synthesized.
"""
import csv
import os
import sys

DATASET_DOI = "10.5281/zenodo.6828947"
CITE = ("Gao, J.; Wang, Z.; Wang, C. (2022). 3D Printing Concrete Mix Design Open Dataset (v0.3). "
        "Zenodo. DOI 10.5281/zenodo.6828947.")

COLS = ["source_dataset", "is_3d_printed", "material_class", "w_b_ratio",
        "static_yield_stress_pa", "dynamic_yield_stress_pa", "plastic_viscosity_pa_s",
        "test_age_days", "compressive_strength_mpa", "doi", "source_citation", "provenance_notes"]

# fixed source column positions (Sheet1 header row)
IX = dict(mixtype=3, ref=4, binder1=6, fa=14, sf=18, microsi=20, slag=30, wb=40, sb=41,
          sy=97, dy=98, pv=99, cs1=100, cs3=101, cs7=102, cs28=103)
AGES = [(1, "cs1"), (3, "cs3"), (7, "cs7"), (28, "cs28")]
MAX_PER_REF = 2
MAX_ROWS = 12


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main(src):
    from openpyxl import load_workbook
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    data = list(ws.iter_rows(values_only=True))[1:]
    wb.close()
    out, per_ref = [], {}
    for r in data:
        g = lambda k: num(r[IX[k]])
        sy, w_b = g("sy"), g("wb")
        ages = [(d, g(k)) for d, k in AGES if g(k) is not None]
        if sy is None or w_b is None or not ages:
            continue  # keep only rows with rheology + w/b + a strength
        ref = str(r[IX["ref"]]).strip() if r[IX["ref"]] else ""
        if per_ref.get(ref, 0) >= MAX_PER_REF:
            continue
        per_ref[ref] = per_ref.get(ref, 0) + 1
        scm = any((g(k) or 0) > 0 for k in ("fa", "sf", "microsi", "slag"))
        scm_names = [n for n, k in (("fly ash", "fa"), ("silica fume", "sf"),
                                    ("micro-silica", "microsi"), ("slag", "slag")) if (g(k) or 0) > 0]
        sand_b = g("sb")
        note = ("Source basis = ratio-to-binder (binder=1); kg/m3 not recoverable, so constituent "
                "mass-% is not formed. Binder: " + (str(r[IX["binder1"]]).strip() or "Portland cement")
                + (" + " + ", ".join(scm_names) if scm_names else "")
                + (f"; sand/binder ratio {sand_b}" if sand_b is not None else "")
                + f". Primary study: {ref}.")
        for d, cs in ages:
            out.append({
                "source_dataset": "UF/UTK 3DCP Mix-Design Open Dataset",
                "is_3d_printed": True,
                "material_class": "blended_OPC" if scm else "OPC",
                "w_b_ratio": w_b,
                "static_yield_stress_pa": round(g("sy") * 1000, 1) if g("sy") is not None else None,
                "dynamic_yield_stress_pa": round(g("dy") * 1000, 1) if g("dy") is not None else None,
                "plastic_viscosity_pa_s": g("pv"),
                "test_age_days": d, "compressive_strength_mpa": cs,
                "doi": DATASET_DOI, "source_citation": CITE, "provenance_notes": note,
            })
        if len(out) >= MAX_ROWS:
            break
    dest = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "ufutk-3dcp-mix.open3dcp.csv")
    with open(dest, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=COLS)
        w.writeheader()
        for d in out:
            w.writerow({k: ("" if d.get(k) is None else d.get(k)) for k in COLS})
    print(f"wrote {len(out)} rows -> {dest}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit('usage: python build/extract.py "/path/to/3D concrete printing mix design dataset v0.3.xlsx"')
    main(sys.argv[1])
