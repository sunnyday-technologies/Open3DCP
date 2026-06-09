#!/usr/bin/env python3
"""Reproduce the TU-Braunschweig "Database of 3D Concrete Printed Buildings" -> Open3DCP excerpt.

This example demonstrates a DIFFERENT slice of the digital twin: the **project / process layer** (who
printed what, where, with which system) rather than mix design or mechanical properties. The source carries
NO mix composition, rheology, or strength data -- and that is precisely the point: not every open 3DCP
dataset is a materials dataset, and a connective schema has to reach the project/process layer too. The
current Open3DCP flat schema is mix-centric, so the building-project fields below have no flat materials
home; they are curated as named project-layer columns + provenance, with NO mix-fidelity score.

Source (CC BY 4.0), NOT re-hosted -- download it from the DOI and run:

    python build/extract.py "/path/to/Database_3D-printed-buildings.xlsx"

    Placzek, G.; Dahlberg, M. "Database of 3D Concrete Printed Buildings"
    Zenodo, doi:10.5281/zenodo.14214812 (CC BY 4.0); paper doi:10.3390/buildings14113410

It selects ten landmark extrusion-printed buildings (2016-2020) spanning consortia, countries, and
fabrication strategies, and writes ../tu-braunschweig-buildings.open3dcp.csv. Values are real.
"""
import csv
import os
import sys

DOI = "10.5281/zenodo.14214812"
CITE = ("Placzek, G.; Dahlberg, M. (2024). Database of 3D Concrete Printed Buildings [Dataset]. "
        "Zenodo. DOI 10.5281/zenodo.14214812.")
# Landmark records (by source ID) spanning years, consortia, countries, and strategies.
SELECT_IDS = [12, 18, 20, 45, 47, 50, 63, 64, 67, 58]

COLS = ["project_name", "year", "country", "consortium", "printer", "fabrication_strategy",
        "printing_strategy", "storeys", "floor_area_m2", "purpose", "is_3d_printed",
        "doi", "source_citation", "provenance_notes"]
SRC = {"id": "ID", "year": "Year", "consortium": "Project Consortium", "project": "Construction Project",
       "city": "City", "country": "Country", "purpose": "Purpose", "storeys": "Storeys",
       "m2": "Square Meter", "fab": "Fabrication Strategy", "print": "Printing Strategy",
       "printer": "Printer", "btype": "Type of Building"}


def clean(v):
    if v is None:
        return None
    return str(v).replace(" ", " ").replace("\xa0", " ").strip() or None


def main(src):
    from openpyxl import load_workbook
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [str(c).strip() if c is not None else "" for c in rows[0]]
    ix = {k: (hdr.index(v) if v in hdr else None) for k, v in SRC.items()}
    by_id = {}
    for r in rows[1:]:
        if ix["id"] is None or r[ix["id"]] is None:
            continue
        try:
            by_id[int(r[ix["id"]])] = r
        except (TypeError, ValueError):
            pass
    out = []
    for sid in SELECT_IDS:
        r = by_id.get(sid)
        if r is None:
            raise SystemExit(f"source ID {sid} not found")
        g = lambda k: clean(r[ix[k]]) if ix[k] is not None else None
        note = (f"Project/process-layer record (no mix/strength data). "
                f"Building type: {g('btype') or 'n/a'}; city: {g('city') or 'n/a'}. "
                f"Source: TU-Braunschweig Database of 3D Concrete Printed Buildings, record ID {sid}.")
        out.append({
            "project_name": g("project"), "year": g("year"), "country": g("country"),
            "consortium": g("consortium"), "printer": g("printer"),
            "fabrication_strategy": g("fab"), "printing_strategy": g("print"),
            "storeys": g("storeys"), "floor_area_m2": g("m2"), "purpose": g("purpose"),
            "is_3d_printed": True, "doi": DOI, "source_citation": CITE, "provenance_notes": note,
        })
    dest = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "tu-braunschweig-buildings.open3dcp.csv")
    with open(dest, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=COLS)
        w.writeheader()
        for d in out:
            w.writerow({k: ("" if d.get(k) is None else d.get(k)) for k in COLS})
    print(f"wrote {len(out)} project-layer records -> {dest}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit('usage: python build/extract.py "/path/to/Database_3D-printed-buildings.xlsx"')
    main(sys.argv[1])
